from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Sheet 1: Resource Pricing ---
ws = wb.active
ws.title = "Resource Pricing"

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E79')
data_font = Font(name='Arial', size=10)
currency_fmt = '$#,##0.000000'
money_fmt = '$#,##0.00'
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

headers = ['Resource', 'SKU', 'Region', 'Pricing Model', 'Unit Price (USD)', 'Unit', 'Hourly Rate (USD)', 'Notes']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

data = [
    ['VM Compute', 'Standard_D2s_v5 Windows', 'indonesiacentral', 'Consumption', 0.200, '1 Hour', 0.200, 'INCL. Windows license; 2 vCPU 8GB'],
    ['VM OS Disk', 'Standard HDD S10 128GB LRS', 'indonesiacentral', 'Consumption', 5.89, '1 Month', None, 'Standard_LRS; 127GB actual'],
    ['SQL Compute (per vCore)', 'Hyperscale Gen5 Serverless (HS_S_Gen5_2)', 'indonesiacentral', 'Consumption', 0.6849, '1 vCore/Hour', 0.6849, '2 vCores max; Serverless auto-pause'],
    ['SQL Storage', 'Hyperscale Data Stored', 'indonesiacentral', 'Consumption', 0.108, '1 GB/Month', None, '$0.108/GB/month'],
    ['SQL Backup (LRS)', 'Hyperscale Backup LRS', 'indonesiacentral', 'Consumption', 0.0864, '1 GB/Month', None, '$0.0864/GB/month'],
    ['Fabric Capacity', 'F2 (2 CUs)', 'indonesiacentral', 'Provisioned', 0.38, '1 Hour', 0.38, '2 CUs x $0.19/CU/hr'],
    ['Fabric - Mirrored DB', 'Mirroring (within F2)', 'indonesiacentral', 'Included in F2', 0, 'CU consumption', 0, 'Consumes CUs from F2'],
    ['Fabric - SQL Endpoint', 'SQL Analytics Endpoint', 'indonesiacentral', 'Included in F2', 0, 'CU consumption', 0, 'Consumes CUs from F2'],
    ['Fabric - Semantic Model', 'Direct Lake on SQL', 'indonesiacentral', 'Included in F2', 0, 'CU consumption', 0, 'Consumes CUs from F2'],
    ['Fabric - Power BI Report', 'FAQ_rpt', 'indonesiacentral', 'Included in F2', 0, 'CU consumption', 0, 'Consumes CUs from F2'],
    ['GPT-5 Input', 'GlobalStandard cap-20', 'eastus2', 'Per Token', 0.0025, '1K tokens', None, 'Est. gpt-4o-1120 rate'],
    ['GPT-5 Output', 'GlobalStandard cap-20', 'eastus2', 'Per Token', 0.01, '1K tokens', None, 'Est. gpt-4o-1120 rate'],
    ['text-embedding-3-small', 'GlobalStandard cap-50', 'eastus2', 'Per Token', 0.000025, '1K tokens', None, 'Confirmed from API'],
    ['AI Services Account', 'S0', 'eastus2', 'Free', 0, 'Base', 0, 'No base charge'],
    ['VNet', 'vnet-lab513-2139d8', 'indonesiacentral', 'Free', 0, 'N/A', 0, 'No charge'],
    ['NSG', 'nsg-lab513-2139d8', 'indonesiacentral', 'Free', 0, 'N/A', 0, 'No charge'],
    ['AI Foundry Project', 'FAQ-Assistant-project', 'eastus2', 'Free', 0, 'N/A', 0, 'Included in AI Services'],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = data_font
        c.border = thin_border
        if col_idx in (5, 7):
            c.number_format = currency_fmt

ws['G3'] = '=E3/730'
ws['G5'] = '=E5/730'
ws['G6'] = '=E6/730'

col_widths = [24, 28, 18, 16, 16, 14, 18, 32]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 2: 8-Hour Simulation ---
ws2 = wb.create_sheet("8-Hour Lab Simulation")

title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
section_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
input_fill = PatternFill('solid', fgColor='FFF2CC')
result_fill = PatternFill('solid', fgColor='D9E2F3')
total_fill = PatternFill('solid', fgColor='1F4E79')
total_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
fabric_fill = PatternFill('solid', fgColor='E2EFDA')

ws2['A1'] = 'Lab513 Cost Simulation - 8 Hour Workshop (incl. Fabric F2)'
ws2['A1'].font = title_font
ws2.merge_cells('A1:F1')

ws2['A3'] = 'SIMULATION PARAMETERS'
ws2['A3'].font = section_font

params = [
    ['Duration (hours)', 8, '', '', ''],
    ['SQL vCores (Hyperscale Gen5)', 2.0, 'vCores', '', 'HS_S_Gen5_2 Serverless, max 2'],
    ['SQL Storage used (GB)', 10, 'GB', '', ''],
    ['Chat interactions', 50, 'conversations', '', 'During 8hr workshop'],
    ['Avg input tokens per chat', 500, 'tokens', '', ''],
    ['Avg output tokens per chat', 300, 'tokens', '', ''],
    ['Embedding calls', 100, 'calls', '', ''],
    ['Avg tokens per embedding', 200, 'tokens', '', ''],
    ['Fabric F2 active', 'YES', '', '', 'F2 provisioned = always-on'],
]

for row_idx, row_data in enumerate(params, 4):
    ws2.cell(row=row_idx, column=1, value=row_data[0]).font = data_font
    c = ws2.cell(row=row_idx, column=2, value=row_data[1])
    c.font = Font(name='Arial', size=10, color='0000FF')
    c.fill = input_fill
    c.border = thin_border
    ws2.cell(row=row_idx, column=3, value=row_data[2]).font = data_font
    ws2.cell(row=row_idx, column=5, value=row_data[4]).font = Font(name='Arial', size=9, italic=True, color='666666')

ws2['A14'] = 'COST BREAKDOWN (8 Hours)'
ws2['A14'].font = section_font
ws2['A14'].border = thin_border

# Headers with Region column
cost_headers = ['Component', 'Region', 'Calculation', 'Cost (USD)']
for col, h in enumerate(cost_headers, 1):
    c = ws2.cell(row=15, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')

rows = [
    ['VM Compute (D2s_v5 Windows)', 'indonesiacentral', '8 hrs x $0.200/hr', '=B4*0.2'],
    ['VM OS Disk (S10 128GB)', 'indonesiacentral', '8 hrs x $5.89/730/hr', '=B4*(5.89/730)'],
    ['SQL Hyperscale Gen5 (HS_S_Gen5_2)', 'indonesiacentral', '8 hrs x 2 vCores x $0.6849', '=B4*B5*0.6849'],
    ['SQL Storage (Hyperscale)', 'indonesiacentral', 'GB x $0.108/mo prorated', '=B6*0.108*(B4/730)'],
    ['Microsoft Fabric F2 Capacity', 'indonesiacentral', '8 hrs x $0.38/hr (2 CUs)', '=B4*0.38'],
    ['  - Mirrored DB + Endpoint + Model + Report', 'indonesiacentral', 'Included in F2', 0],
    ['GPT-5 Input Tokens', 'eastus2', 'chats x tokens x $0.0025/1K', '=B7*B8/1000*0.0025'],
    ['GPT-5 Output Tokens', 'eastus2', 'chats x tokens x $0.01/1K', '=B7*B9/1000*0.01'],
    ['Embedding Tokens', 'eastus2', 'calls x tokens x $0.000025/1K', '=B10*B11/1000*0.000025'],
    ['VNet + NSG + AI Project', 'both', 'Free', 0],
]

for i, row_data in enumerate(rows, 16):
    ws2.cell(row=i, column=1, value=row_data[0]).font = data_font
    ws2.cell(row=i, column=1).border = thin_border
    ws2.cell(row=i, column=2, value=row_data[1]).font = Font(name='Arial', size=9, color='666666')
    ws2.cell(row=i, column=2).border = thin_border
    ws2.cell(row=i, column=3, value=row_data[2]).font = data_font
    ws2.cell(row=i, column=3).border = thin_border
    c = ws2.cell(row=i, column=4, value=row_data[3])
    c.number_format = money_fmt
    c.border = thin_border
    if 'Fabric' in str(row_data[0]):
        for col in range(1, 5):
            ws2.cell(row=i, column=col).fill = fabric_fill

# TOTAL row
tr = 26
ws2.cell(row=tr, column=1, value='TOTAL COST (8 HOURS)').font = total_font
ws2.cell(row=tr, column=1).fill = total_fill
ws2.cell(row=tr, column=2).fill = total_fill
ws2.cell(row=tr, column=3).fill = total_fill
ws2.cell(row=tr, column=4, value='=SUM(D16:D25)').font = total_font
ws2.cell(row=tr, column=4).fill = total_fill
ws2.cell(row=tr, column=4).number_format = money_fmt
ws2.cell(row=tr, column=4).border = thin_border

# Explanation
ws2.cell(row=28, column=1, value='HOW TO READ:').font = section_font
ws2.cell(row=29, column=1, value='Total Cost (above) = total biaya selama DURASI yang dipilih (8 jam)').font = data_font
ws2.cell(row=30, column=1, value='Per Hour = Total Cost / Duration').font = data_font
ws2['D30'] = '=D26/B4'
ws2['D30'].number_format = money_fmt
ws2['D30'].border = thin_border
ws2.cell(row=31, column=1, value='Monthly projection (730 hrs always-on)').font = data_font
ws2['D31'] = '=D30*730'
ws2['D31'].number_format = money_fmt
ws2['D31'].border = thin_border

# DEPLOY REGIONS
ws2.cell(row=33, column=1, value='DEPLOY REGIONS:').font = section_font
ws2.cell(row=34, column=1, value='indonesiacentral').font = Font(name='Arial', size=10, bold=True)
ws2.cell(row=34, column=2, value='VM, Disk, SQL Hyperscale, Fabric F2, VNet, NSG').font = data_font
ws2.cell(row=35, column=1, value='eastus2').font = Font(name='Arial', size=10, bold=True)
ws2.cell(row=35, column=2, value='AI Foundry, GPT-5, text-embedding-3-small').font = data_font

# Disclaimer
ws2.cell(row=37, column=1, value='DISCLAIMER:').font = Font(name='Arial', bold=True, size=10, color='C00000')
ws2.cell(row=38, column=1, value='Harga diperoleh melalui Azure Retail Prices API (https://prices.azure.com/api/retail/prices)').font = Font(name='Arial', size=9)
ws2.cell(row=39, column=1, value='Harga dapat berubah sewaktu-waktu. Periksa Azure Portal untuk harga terkini.').font = Font(name='Arial', size=9)
ws2.cell(row=40, column=1, value='VM D2s_v5 Windows $0.200/hr SUDAH TERMASUK lisensi Windows Server.').font = Font(name='Arial', size=9, bold=True)
ws2.cell(row=41, column=1, value='GPT-5: estimasi dari gpt-4o-1120 GlobalStandard (model belum tersedia di retail API).').font = Font(name='Arial', size=9)

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 28

# --- Sheet 3: Hourly Breakdown ---
ws3 = wb.create_sheet("Hourly Breakdown")

ws3['A1'] = 'Hour-by-Hour Cost Accumulation (with Fabric F2)'
ws3['A1'].font = title_font
ws3.merge_cells('A1:H1')

h3 = ['Hour', 'VM ($)', 'Disk ($)', 'SQL ($)', 'Fabric ($)', 'AI Tokens ($)', 'Hourly Total ($)', 'Running Total ($)']
for col, h in enumerate(h3, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')

for hr in range(1, 9):
    row = hr + 3
    ws3.cell(row=row, column=1, value=hr).border = thin_border
    ws3.cell(row=row, column=2, value=0.200).number_format = money_fmt
    ws3.cell(row=row, column=2).border = thin_border
    ws3[f'C{row}'] = '=5.89/730'
    ws3[f'C{row}'].number_format = '$#,##0.0000'
    ws3[f'C{row}'].border = thin_border
    ws3.cell(row=row, column=4, value=0.6849).number_format = money_fmt
    ws3.cell(row=row, column=4).border = thin_border
    ws3.cell(row=row, column=5, value=0.38).number_format = money_fmt
    ws3.cell(row=row, column=5).border = thin_border
    ws3.cell(row=row, column=5).fill = fabric_fill
    ws3[f'F{row}'] = '=(50*500/1000*0.0025+50*300/1000*0.01+100*200/1000*0.000025)/8'
    ws3[f'F{row}'].number_format = '$#,##0.0000'
    ws3[f'F{row}'].border = thin_border
    ws3[f'G{row}'] = f'=B{row}+C{row}+D{row}+E{row}+F{row}'
    ws3[f'G{row}'].number_format = money_fmt
    ws3[f'G{row}'].border = thin_border
    if hr == 1:
        ws3[f'H{row}'] = f'=G{row}'
    else:
        ws3[f'H{row}'] = f'=H{row-1}+G{row}'
    ws3[f'H{row}'].number_format = money_fmt
    ws3[f'H{row}'].font = Font(name='Arial', bold=True)
    ws3[f'H{row}'].fill = result_fill
    ws3[f'H{row}'].border = thin_border

ws3.cell(row=12, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=11)
ws3['H12'] = '=H11'
ws3['H12'].font = total_font
ws3['H12'].fill = total_fill
ws3['H12'].number_format = money_fmt
ws3['H12'].border = thin_border

ws3['A14'] = 'Regions: VM/Disk/SQL/Fabric = indonesiacentral | AI Tokens = eastus2'
ws3['A14'].font = Font(name='Arial', size=9, italic=True, color='666666')

for col in range(1, 9):
    ws3.column_dimensions[get_column_letter(col)].width = 14

# --- Sheet 4: Fabric Detail ---
ws4 = wb.create_sheet("Fabric Resources")

ws4['A1'] = 'Microsoft Fabric Resources (Lab Exercise 5)'
ws4['A1'].font = title_font
ws4.merge_cells('A1:D1')

ws4['A3'] = 'Fabric Capacity SKU'
ws4['A3'].font = section_font
ws4['B3'] = 'F2 (2 Capacity Units)'
ws4['A4'] = 'Region'
ws4['B4'] = 'indonesiacentral'
ws4['A5'] = 'CU Rate'
ws4['B5'] = '$0.19/CU/hour'
ws4['A6'] = 'F2 Hourly Cost'
ws4['B6'] = '=2*0.19'
ws4['B6'].number_format = money_fmt
ws4['A7'] = 'F2 Monthly Cost (730hr)'
ws4['B7'] = '=B6*730'
ws4['B7'].number_format = money_fmt

ws4['A9'] = 'FABRIC ITEMS'
ws4['A9'].font = section_font

fab_headers = ['Item', 'Type', 'Region', 'Notes']
for col, h in enumerate(fab_headers, 1):
    c = ws4.cell(row=10, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

fabric_items = [
    ['Workspace2139d8', 'Workspace', 'indonesiacentral', 'Named per LAB_INSTANCE_ID'],
    ['faq-ai-assistant-db', 'Mirrored Azure SQL DB', 'indonesiacentral', 'Mirrors dbo.FAQ_Content to OneLake'],
    ['SQL Analytics Endpoint', 'Auto-created', 'indonesiacentral', 'Query mirrored Delta tables'],
    ['FAQ_Content', 'Semantic Model (Direct Lake)', 'indonesiacentral', 'Storage: Direct Lake on SQL'],
    ['FAQ_rpt', 'Power BI Report', 'indonesiacentral', 'Created via Copilot in Fabric'],
]

for row_idx, row_data in enumerate(fabric_items, 11):
    for col_idx, val in enumerate(row_data, 1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        c.font = data_font
        c.border = thin_border

ws4['A17'] = 'ARCHITECTURE'
ws4['A17'].font = section_font
ws4['A18'] = 'Azure SQL Hyperscale -> Fabric Mirroring -> OneLake (Delta) -> SQL Endpoint -> Semantic Model -> Power BI'
ws4['A18'].font = Font(name='Consolas', size=9)

ws4.column_dimensions['A'].width = 24
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 34

# Save
output_path = r'C:\Lab\envlab513\cost-calculator-8hr.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
